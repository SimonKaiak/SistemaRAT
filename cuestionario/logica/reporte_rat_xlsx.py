"""
reporte_rat_xlsx.py
-------------------
Vista para la descarga de datos RAT en formato Excel (.xlsx).

descargar_datos_xlsx(request)
    Genera un Excel con una fila por trabajador y una columna
    por pregunta RAT, incluyendo datos del colaborador.

    Control de acceso:
    - Superusuario: filtra por empresa vía query param empresa_id.
    - Coordinador: solo su empresa.
    - Trabajador regular: redirige a index.
"""

from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from cuestionario.models import (
    Trabajador, Empresa, RATRespuestas, InstrumentoEmpresa
)
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def descargar_datos_xlsx(request):

    # ── Control de acceso ──────────────────────
    if not request.user.is_superuser:
        try:
            trabajador_actual = Trabajador.objects.get(user=request.user)
            if not trabajador_actual.es_coordinador:
                return redirect('index')
            empresa_obj = trabajador_actual.empresa
        except Trabajador.DoesNotExist:
            return redirect('index')
    else:
        empresa_id = request.GET.get('empresa_id')
        empresa_obj = None
        if empresa_id:
            try:
                empresa_obj = Empresa.objects.get(id_empresa=empresa_id)
            except Empresa.DoesNotExist:
                return HttpResponse("Empresa no encontrada.", status=404)

    if not empresa_obj:
        return HttpResponse("No se especificó empresa.", status=400)

    # ── Instrumento RAT activo ─────────────────
    ie = InstrumentoEmpresa.objects.filter(
        empresa=empresa_obj, habilitado=True, instrumento__tipo='rat'
    ).first()
    if not ie:
        return HttpResponse("No hay instrumento RAT habilitado para esta empresa.", status=404)

    preguntas = ie.preguntas.exclude(
        tipo='texto', actividad_tratamiento__startswith='Presentación'
    ).exclude(
        actividad_tratamiento__icontains='Fuente de la cual provienen'
    ).order_by('id_rat_pregunta')

    # ── Trabajadores con RAT listo ─────────────
    total_preguntas = preguntas.count()
    trabajadores_empresa = Trabajador.objects.filter(empresa=empresa_obj)
    trabajadores_listos = []
    for t in trabajadores_empresa:
        respondidas = RATRespuestas.objects.filter(
            trabajador=t, pregunta__instrumento_empresa=ie
        ).exclude(
            pregunta__tipo='texto',
            pregunta__actividad_tratamiento__startswith='Presentación'
        ).exclude(
            pregunta__actividad_tratamiento__icontains='Fuente de la cual provienen'
        ).values('pregunta').distinct().count()
        if total_preguntas > 0 and respondidas >= total_preguntas:
            trabajadores_listos.append(t)

    # ── Construir Excel ────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos RAT"

    headers = ['Colaborador', 'Cargo', 'Nivel', 'Departamento'] + \
              [f"P{i+1}: {p.actividad_tratamiento[:40]}" for i, p in enumerate(preguntas)]
    ws.append(headers)

    for t in trabajadores_listos:
        fila = [
            f"{t.nombre} {t.apellido_paterno} {t.apellido_materno}",
            t.cargo.nombre_cargo,
            t.nivel_jerarquico.nombre_nivel_jerarquico if t.nivel_jerarquico else '',
            t.departamento.nombre_departamento if t.departamento else '',
        ]
        for pregunta in preguntas:
            resp = RATRespuestas.objects.filter(trabajador=t, pregunta=pregunta).first()
            fila.append(resp.respuesta if resp else '')
        ws.append(fila)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 30

    # ── Respuesta ──────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = (
        f"datos_rat_{empresa_obj.nombre_empresa.replace(' ', '_')}"
        f"_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response